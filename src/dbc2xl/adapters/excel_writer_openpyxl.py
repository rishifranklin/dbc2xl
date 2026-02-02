from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..domain.models import AttributeExport, DbcExport, MessageExport, NodeExport, SignalExport
from ..utils.formatting import safe_str, stringify_choices

logger = logging.getLogger(__name__)


_HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=False)
_CELL_ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)


def _autosize_columns(ws, max_width: int = 60) -> None:
    # openpyxl has no true autosize; we estimate based on content length
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            s = "" if value is None else str(value)
            dims[idx] = max(dims.get(idx, 0), len(s))
    for idx, width in dims.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), max_width)


def _apply_table_header(ws, header_row: int = 1) -> None:
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _set_wrap_for_columns(ws, wrap_cols: Sequence[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        for col_idx in wrap_cols:
            if col_idx <= len(row):
                row[col_idx - 1].alignment = _CELL_ALIGN_WRAP


def _write_sheet(ws, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    _apply_table_header(ws, 1)
    _autosize_columns(ws)


def _bool_str(v: object) -> str:
    if v is None:
        return ""
    return "Yes" if bool(v) else "No"


@dataclass(frozen=True)
class OpenpyxlExcelWriter:
    def write(self, export: DbcExport, xlsx_path: str) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        self._write_messages(wb, export.messages)
        self._write_signals(wb, export.signals)
        self._write_nodes(wb, export.nodes)
        self._write_attributes(wb, export.attributes)
        self._write_value_tables(wb, export.signals)

        wb.save(xlsx_path)

    def _write_messages(self, wb: Workbook, messages: Sequence[MessageExport]) -> None:
        ws = wb.create_sheet("Messages")
        headers = [
            "Message Name",
            "Frame ID (Hex)",
            "Frame ID (Dec)",
            "Extended Frame",
            "DLC/Length",
            "Cycle Time (ms)",
            "Senders (Origin Nodes)",
            "Comment",
            "Attributes (Key=Value; ...)",
        ]

        def rows():
            for m in messages:
                attrs = "; ".join([f"{k}={safe_str(v)}" for k, v in sorted(m.attributes.items())])
                yield [
                    m.name,
                    m.frame_id_hex,
                    m.frame_id,
                    _bool_str(m.is_extended_frame),
                    m.length,
                    "" if m.cycle_time_ms is None else m.cycle_time_ms,
                    ", ".join(m.senders),
                    safe_str(m.comment),
                    attrs,
                ]

        _write_sheet(ws, headers, rows())
        # Wrap comment + attributes
        _set_wrap_for_columns(ws, wrap_cols=[8, 9])

    def _write_signals(self, wb: Workbook, signals: Sequence[SignalExport]) -> None:
        ws = wb.create_sheet("Signals")
        headers = [
            "Message Name",
            "Msg Frame ID (Hex)",
            "Msg Frame ID (Dec)",
            "Signal Name",
            "Start Bit",
            "Length (bits)",
            "Byte Order",
            "Signed",
            "Float",
            "Factor",
            "Offset",
            "Min",
            "Max",
            "Unit",
            "Receivers",
            "Multiplexer",
            "Mux IDs",
            "Mux Signal",
            "Choices / Value Table",
            "Comment",
            "Attributes (Key=Value; ...)",
        ]

        def rows():
            for s in signals:
                attrs = "; ".join([f"{k}={safe_str(v)}" for k, v in sorted(s.attributes.items())])
                choices_str = stringify_choices(s.choices) if s.choices else ""
                yield [
                    s.message_name,
                    s.message_frame_id_hex,
                    s.message_frame_id,
                    s.name,
                    s.start,
                    s.length,
                    safe_str(s.byte_order),
                    _bool_str(s.is_signed),
                    _bool_str(s.is_float),
                    "" if s.factor is None else s.factor,
                    "" if s.offset is None else s.offset,
                    "" if s.minimum is None else s.minimum,
                    "" if s.maximum is None else s.maximum,
                    safe_str(s.unit),
                    ", ".join(s.receivers),
                    _bool_str(s.is_multiplexer),
                    "" if not s.multiplexer_ids else ", ".join([str(x) for x in s.multiplexer_ids]),
                    safe_str(s.multiplexer_signal),
                    choices_str,
                    safe_str(s.comment),
                    attrs,
                ]

        _write_sheet(ws, headers, rows())
        # Wrap value table + comment + attrs
        _set_wrap_for_columns(ws, wrap_cols=[19, 20, 21])

    def _write_nodes(self, wb: Workbook, nodes: Sequence[NodeExport]) -> None:
        ws = wb.create_sheet("Nodes")
        headers = ["Node Name", "Comment", "Attributes (Key=Value; ...)"]

        def rows():
            for n in nodes:
                attrs = "; ".join([f"{k}={safe_str(v)}" for k, v in sorted(n.attributes.items())])
                yield [n.name, safe_str(n.comment), attrs]

        _write_sheet(ws, headers, rows())
        _set_wrap_for_columns(ws, wrap_cols=[2, 3])

    def _write_attributes(self, wb: Workbook, attrs: Sequence[AttributeExport]) -> None:
        ws = wb.create_sheet("Attributes")
        headers = ["Scope", "Owner", "Key", "Value"]

        def rows():
            for a in attrs:
                yield [a.scope, a.owner, a.key, safe_str(a.value)]

        _write_sheet(ws, headers, rows())

    def _write_value_tables(self, wb: Workbook, signals: Sequence[SignalExport]) -> None:
        ws = wb.create_sheet("ValueTables")
        headers = ["Message", "Signal", "Frame ID (Hex)", "Value", "Text"]

        def rows():
            for s in signals:
                if not s.choices:
                    continue
                for k, v in sorted(s.choices.items(), key=lambda kv: kv[0]):
                    yield [s.message_name, s.name, s.message_frame_id_hex, k, v]

        _write_sheet(ws, headers, rows())
